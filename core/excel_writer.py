import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from config import TABLE_SCHEMAS


# ── Color palette ──────────────────────────────────────────────────────────────
HEADER_BG    = "1F3864"   # dark navy
HEADER_FG    = "FFFFFF"
ALT_ROW_BG   = "EEF2FF"   # light blue-grey
FK_COLOR     = "004B8D"   # blue for FK columns
PK_COLOR     = "7B2D8B"   # purple for PK columns
BORDER_COLOR = "C5CAE9"


def _header_style(cell, is_pk=False, is_fk=False):
    cell.font = Font(
        bold=True, color=HEADER_FG, name="Arial", size=10
    )
    cell.fill = PatternFill("solid", start_color=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=BORDER_COLOR)
    cell.border = Border(left=thin, right=thin, bottom=thin, top=thin)


def _data_style(cell, row_idx: int):
    thin = Side(style="thin", color=BORDER_COLOR)
    cell.border = Border(left=thin, right=thin, bottom=thin, top=thin)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if row_idx % 2 == 0:
        cell.fill = PatternFill("solid", start_color=ALT_ROW_BG)
    cell.font = Font(name="Arial", size=9)


def _auto_width(sheet, max_width=50):
    for col_cells in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def _write_sheet(ws, columns: list[str], rows: list[dict], sheet_title: str):
    """Write headers + data rows into a worksheet."""
    ws.title = sheet_title[:31]  # Excel sheet name limit

    # Freeze header row
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    # Write headers
    pk_cols = {c for c in columns if c.endswith("_id") and not c.startswith("candidate")}
    fk_cols = {"candidate_id"}

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
        is_pk = col_name in pk_cols
        is_fk = col_name in fk_cols
        _header_style(cell, is_pk=is_pk, is_fk=is_fk)

    # Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = row_data.get(col_name)
            if val is None:
                val = ""
            elif isinstance(val, list):
                val = "; ".join(str(v) for v in val)
            elif isinstance(val, bool):
                val = "Yes" if val else "No"
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _data_style(cell, row_idx)

    _auto_width(ws)


def write_excel(all_candidates: list[dict], output_path: str):
    """
    Main entry point. Takes list of fully processed candidate dicts,
    writes one Excel file with one sheet per table.
    """
    wb = openpyxl.Workbook()

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Collect rows per table across all candidates
    table_rows: dict[str, list[dict]] = {t: [] for t in TABLE_SCHEMAS}

    for cand in all_candidates:
        cid = cand.get("candidate_id", "unknown")

        # personal_info — one row per candidate
        pi = dict(cand.get("personal_info", {}))
        pi["candidate_id"] = cid
        pi["source_pdf"] = cand.get("source_pdf", "")
        table_rows["personal_info"].append(pi)

        # All list tables
        # BUG FIX 6: supervision is stored as a dict (aggregate counts), not a
        # list of records. pipeline.py sets it to a dict like
        # {"ms_main_supervisor": 2, "student_names": [...], ...}.
        # Treating it as a list caused silent data loss. Convert it to a single
        # flat row before appending so it appears correctly in the Excel sheet.
        raw_sup = cand.get("supervision")
        if isinstance(raw_sup, dict):
            sup_row = {
                "record_id": f"sup_{cid}_01",
                "candidate_id": cid,
                "student_name": "; ".join(raw_sup.get("student_names") or []),
                "degree_supervised": "MS/PhD",
                "supervision_role": (
                    f"MS-Main:{raw_sup.get('ms_main_supervisor',0)} "
                    f"MS-Co:{raw_sup.get('ms_co_supervisor',0)} "
                    f"PhD-Main:{raw_sup.get('phd_main_supervisor',0)} "
                    f"PhD-Co:{raw_sup.get('phd_co_supervisor',0)}"
                ),
                "year_graduated": None,
                "joint_publication": raw_sup.get("publications_with_students", 0),
            }
            table_rows["supervision"].append(sup_row)

        for table_name in TABLE_SCHEMAS:
            if table_name in ("personal_info", "supervision"):
                continue
            records = cand.get(table_name, [])
            for seq, rec in enumerate(records, 1):
                row = dict(rec)
                row["candidate_id"] = cid
                prefix = table_name[:3]
                row["record_id"] = f"{prefix}_{cid}_{seq:02d}"
                table_rows[table_name].append(row)

    # Write each table as a sheet
    sheet_display_names = {
        "personal_info":  "📋 Personal Info",
        "education":      "🎓 Education",
        "experience":     "💼 Experience",
        "publications":   "📄 Publications",
        "skills":         "🛠 Skills",
        "patents":        "🔬 Patents",
        "books":          "📚 Books",
        "awards":         "🏆 Awards",
        "references":     "👥 References",
    }

    for table_name, columns in TABLE_SCHEMAS.items():
        ws = wb.create_sheet(title=sheet_display_names.get(table_name, table_name))

        # Build column list: record_id first, candidate_id second, then the rest
        ordered_cols = []
        if "record_id" in columns:
            ordered_cols.append("record_id")
        ordered_cols.append("candidate_id")
        ordered_cols += [c for c in columns if c not in ("record_id", "candidate_id")]

        _write_sheet(ws, ordered_cols, table_rows[table_name], table_name)

    # Add a README / data dictionary sheet
    _write_readme_sheet(wb)

    wb.save(output_path)
    print(f"  ✅ Excel saved → {output_path}")
    return output_path


def _write_readme_sheet(wb: openpyxl.Workbook):
    ws = wb.create_sheet(title="📖 Data Dictionary", index=0)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 60

    headers = ["Table", "Field", "Description"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        _header_style(cell)

    descriptions = {
        "personal_info": {
            "candidate_id": "Primary key — auto-generated (cand_001…)",
            "full_name": "Candidate's full name",
            "father_guardian_name": "Father or guardian name",
            "current_salary_pkr": "Normalized to full PKR (e.g. 200,000)",
            "expected_salary_pkr": "Normalized to full PKR",
            "apply_date": "Date application was submitted",
            "post_applied_for": "Job title applied for",
            "source_pdf": "Source PDF filename",
        },
        "education": {
            "record_id": "edu_cand_001_01 — table prefix + candidate + seq",
            "candidate_id": "FK → personal_info.candidate_id",
            "degree_level": "SSC | HSSC | Bachelor | Master | PhD | Other",
            "grade_type": "GPA | Percentage | Grade",
        },
        "experience": {
            "is_current": "Yes if end_date is Present",
            "duration_months": "Computed from start/end dates",
            "overlap_flag": "Filled by analysis module — job/study overlap",
            "gap_before_months": "Gap since previous role (analysis module)",
        },
        "publications": {
            "pub_type": "Journal | Conference | Workshop | Other",
            "impact_factor_claimed": "As stated on CV — NOT verified externally",
            "authorship_role": "First | Corresponding | Co-author",
        },
        "skills": {
            "evidence_level": "Filled by skill-alignment module later",
        },
    }

    row = 2
    for table, fields in descriptions.items():
        for field, desc in fields.items():
            ws.cell(row=row, column=1, value=table).font = Font(name="Arial", size=9, bold=True)
            ws.cell(row=row, column=2, value=field).font = Font(name="Arial", size=9, color=FK_COLOR)
            ws.cell(row=row, column=3, value=desc).font = Font(name="Arial", size=9)
            for c in range(1, 4):
                thin = Side(style="thin", color=BORDER_COLOR)
                ws.cell(row=row, column=c).border = Border(
                    left=thin, right=thin, top=thin, bottom=thin
                )
            row += 1